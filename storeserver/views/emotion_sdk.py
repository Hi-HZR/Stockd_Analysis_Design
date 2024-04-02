from django.shortcuts import render
from openpyxl.reader.excel import load_workbook
import aip
from django.http import HttpResponse
from django.shortcuts import render

def emotion_sdk(my_text):
    app_id = '57118734'
    api_key = 'MhsCMaKfJX9JyopuyviPbiK5'
    secret_key = 'RGKHayxmxEfKAN2tqhrZsiCmoBLEQUs3'
    my_nlp = aip.nlp.AipNlp(app_id, api_key, secret_key)

    result = my_nlp.sentimentClassify(my_text)
    value = result.get("items")
    positive_prob = value[0]["positive_prob"]
    negative_prob = value[0]["negative_prob"]
    return positive_prob, negative_prob


# # def emotion_txt(request):
# workbook = load_workbook("D:/PythonProject/djangoProject/xueqiu.xlsx")
# # 选择活动的工作表
# sheet = workbook.active
# total_positive_prob = 0
# total_negative_prob = 0
# comment_count = 0
# for row in sheet.iter_rows(min_row=2):
#     my_text = 'row[3].value'
#     positive_prob, negative_prob = emotion_sdk(my_text)
#     if positive_prob is not None and negative_prob is not None:
#         total_positive_prob += positive_prob
#         total_negative_prob += negative_prob
#         comment_count += 1
# # Calculate average only if there are analyzed comments
# if comment_count > 0:
#     average_positive_prob = total_positive_prob / comment_count
#     average_negative_prob = total_negative_prob / comment_count
#     print("Average positive probability:", average_positive_prob)
#     print("Average negative probability:", average_negative_prob)
# else:
#     print("No valid sentiment scores found for comments.")