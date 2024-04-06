import pandas
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl.reader.excel import load_workbook
from storeserver.models import UserInfo, StorePoint
from storeserver.views.crawl_comment import crawl_comment
from storeserver.views.crawl_price import crawl_price
from datetime import datetime


def upload_file(request):
    result = crawl_comment(request)
    d = pandas.DataFrame(result)
    d.to_excel('xueqiu.xlsx')
    # 上传excel文件,并获取
    file_object = request.FILES.get("excel")
    print(type(file_object))

    # 通过openpyxl打开文件
    exc = load_workbook("D:/PythonProject/djangoProject/xueqiu.xlsx")
    sheet = exc.worksheets[0]
    my_id = 0
    # 循环输入数据
    for row in sheet.iter_rows(min_row=2):
        text_uid = row[1].value
        text_name = row[2].value
        text_comment = row[3].value
        # 原始日期字符串
        date_string = row[4].value

        # 指定年份
        year = 2024

        # 将日期字符串转换为datetime对象
        date_obj = datetime.strptime(date_string, "%m-%d %H:%M")

        # 将年份添加到datetime对象
        date_obj = date_obj.replace(year=year)

        # 将datetime对象转换为YYYY-MM-DD格式的字符串
        text_time = date_obj.strftime("%Y-%m-%d")
        my_id += 1
        try:
            UserInfo.objects.create(id=my_id, name=text_name, uid=text_uid, comment=text_comment,time=text_time)
        except UnicodeEncodeError as e:
            # 如果出现UnicodeEncodeError，则不添加到content_list
            print(f"Error encoding string: {e}")
    return HttpResponse('success')


def upload_all(request):
    upload_file(request)
    crawl_price(request)
    exc_2 = load_workbook("D:/PythonProject/djangoProject/point.xlsx")
    sheet_2 = exc_2.worksheets[0]
    my_id = 0
    # 循环输入数据
    for row in sheet_2.iter_rows(min_row=2):
        text_point = row[1].value
        text_time = row[2].value
        my_id += 1
        try:
            StorePoint.objects.create(id=my_id, point=text_point, time=text_time)
        except UnicodeEncodeError as e:
            # 如果出现UnicodeEncodeError，则不添加到content_list
            print(f"Error encoding string: {e}")
    return render(request,'upload.html')
