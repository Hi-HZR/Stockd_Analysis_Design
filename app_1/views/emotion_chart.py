from django.shortcuts import render
from openpyxl.reader.excel import load_workbook
from app_1.views.emotion_sdk import emotion_sdk


def calculate_average_probs(workbook_path):
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    total_positive_prob = 0
    total_negative_prob = 0
    comment_count = 0
    for row in sheet.iter_rows(min_row=2):
        my_text = row[3].value
        positive_prob, negative_prob = emotion_sdk(my_text)
        if positive_prob is not None and negative_prob is not None:
            total_positive_prob += positive_prob
            total_negative_prob += negative_prob
            comment_count += 1
    if comment_count > 0:

        average_positive_prob = round(total_positive_prob / comment_count, 2)
        average_negative_prob = round(total_negative_prob / comment_count, 2)

        return average_positive_prob, average_negative_prob
    else:
        return None, None


def emotion_chart(request):
    average_positive_prob, average_negative_prob = calculate_average_probs("D:/PythonProject/djangoProject/xueqiu.xlsx")
    if average_positive_prob is not None and average_negative_prob is not None:
        print("评论积极评平均概率：", average_positive_prob)
        print("评论消极评平均概率：", average_negative_prob)
        return render(request, 'emotion_chart.html', {
            'average_positive_prob': average_positive_prob,
            'average_negative_prob': average_negative_prob
        })
    else:
        # 处理没有有效情感分数的情况
        pass
