from django.shortcuts import redirect
from openpyxl import load_workbook
from app_1.models import UserInfo


def file_upload(request):
    # 上传excel文件,并获取
    file_object = request.FILES.get("excel")
    print(type(file_object))

    # 通过openpyxl打开文件
    exc = load_workbook("D:/PythonProject/djangoProject/app_1/views/xueqiu.xlsx")
    sheet = exc.worksheets[0]

    # 循环输出数据
    for row in sheet.iter_rows(min_row=2):
        text_uid = row[1].value
        text_name = row[2].value
        text_comment = row[3].value
        UserInfo.objects.create(name=text_name,uid=text_uid,comment=text_comment)
    print("成功爬取数据，并且上传到数据库")
    return redirect('http://127.0.0.1:8000/info')