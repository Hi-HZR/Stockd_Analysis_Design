import pandas
from django.shortcuts import redirect
from openpyxl.reader.excel import load_workbook
from app_1.models import UserInfo
from app_1.views.crawl_comment import crawl_file
from app_1.views.crawl_price import crawl_price


def upload_file(request):
    result_comment = crawl_file(request)
    d = pandas.DataFrame(result_comment)
    d.to_excel('xueqiu.xlsx')

    # result_price = crawl_price(request)
    # d = pandas.DataFrame(result_price)
    # d.to_excel('price.xlsx')

    # 上传excel文件,并获取
    file_object = request.FILES.get("excel")
    print(type(file_object))

    # 通过openpyxl打开文件
    exc = load_workbook("D:/PythonProject/djangoProject/xueqiu.xlsx")
    sheet = exc.worksheets[0]
    my_id = 0
    # 循环输入数据
    for row in sheet.iter_rows(min_row=2):
        text_uid = row[1].value
        text_name = row[2].value
        text_comment = row[3].value
        my_id += 1
        try:
            UserInfo.objects.create(id=my_id, name=text_name, uid=text_uid, comment=text_comment)
        except UnicodeEncodeError as e:
            # 如果出现UnicodeEncodeError，则不添加到content_list
            print(f"Error encoding string: {e}")

    result_price = crawl_price(request)
    d = pandas.DataFrame(result_price)
    d.to_excel('price.xlsx')

    return redirect('http://127.0.0.1:8000/info')


# def upload_price(request):
#     result_price = crawl_price(request)
#     d = pandas.DataFrame(result_price)
#     d.to_excel('price.xlsx')
